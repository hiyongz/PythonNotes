from openpyxl import load_workbook
workbook = load_workbook(filename = "D:/test.xlsx")
sheet = workbook.active
 
# 按行获取值
for i in sheet.iter_rows(min_row=2, max_row=3, min_col=3,max_col=3):
    for j in i:
        print(j.value)

