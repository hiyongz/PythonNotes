# 确保已安装pandas和openpyxl
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import (
    LineChart,
    Reference,
)

# 简化起见，查询期间设置为年初至今天，Q4或12月适用，年初及其他情况需自行修改
# today = time.strftime('%Y-%m-%d')
# start_day = time.strftime('%Y') + '-01-01'

# 用pandas读取获取网页数据
# url = f'http://www.safe.gov.cn/AppStructured/hlw/RMBQuery.do?startDate={start_day}&endDate={today}&queryYN=true'
# df = pd.read_html(url, encoding='utf-8')[3]

df = pd.read_excel("人民币汇率中间价.xls")

# # 文件保存，先将df存入excel，防止替换后日期列改变
save_dir = f'人民币汇率中间价.xlsx'
writer = pd.ExcelWriter(save_dir)
# df.to_excel(writer, '全部', index=None)

# 删除数据中的‘日’信息，仅截取年月信息以便按月汇总
df['日期'] = df['日期'].str.slice(0, 7)

# 获取指定币种的平均值
df1 = df.groupby(['日期'])[['美元', '欧元', '日元']].mean()

# 设置为各自常用汇率表达形式
df1['美元'] = round(df1['美元'] / 100, 2)  # 1美元兑n人民币
df1['欧元'] = round(df1['欧元'] / 100, 2)  # 1欧元兑n人民币
df1['日元'] = round(100 / df1['日元'], 2)  # 1人民币兑n日元

df1 = df1.reset_index()

# 将汇总表存入excel并保存
df1.to_excel(writer, '图表', index=None)
writer.save()

# 用openpyxl打开文件，修饰格式并创建图表
wb = openpyxl.load_workbook(save_dir)  # 加载既存工作表

# 设置各工作表格式
names = wb.sheetnames
for name in names:
    ws = wb[name]
    irow = ws.max_row
    icol = ws.max_column
    ws.views.sheetView[0].showGridLines = False  # 设置不显示网格线
    ws.views.sheetView[0].zoomScale = 85  # 设置默认缩放比例
    # 仅对日期列设置宽度
    ws.column_dimensions['A'].width = 12
    # 设置使用区域居中显示
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for r in range(1, irow + 1):
        for c in range(1, icol + 1):
            ws.cell(r, c).alignment = align
    # 冻结拆分窗格，基点设为B2单元格
    ws.freeze_panes = 'B2'

# 将’图表‘sheet设置为活动工作簿，同时取消’全部‘sheet的选中状态
wb.active = wb['图表']
# wb['全部'].views.sheetView[0].tabSelected = False

# 以下为创建折线图的内容，分日元、美元和欧元两段，每段代码基本相同，用不到可删掉其中一段

# 创建汇率折线图---日元
data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=13)
c1 = LineChart()
c1.title = "日元汇率推移图"
c1.style = 12
# 设置为不显示图例，不设置时默认显示图例
c1.legend = None
# 图例的位置可以通过设置其位置来控制：右、左、上、下和右上分别为r、l、t、b和tr。默认值为r
# c2.legend.position = 'tr'
# 设置纵轴名称
c1.y_axis.title = "汇率"
c1.y_axis.crossAx = 500
# 设置纵轴最大最小值和步长
c1.y_axis.scaling.min = 15
c1.y_axis.scaling.max = 30
c1.y_axis.majorUnit = 1
c1.x_axis = DateAxis(crossAx=100)
# c1.x_axis.title = "月"

# 自定义图表的宽度和高度，不设置时默认为15*7.5cm
c1.width = 25
c1.height = 10

c1.dLbls = DataLabelList()
# c2.dLbls.showCatName = True  # 数据标签显示x坐标轴标签
c1.dLbls.showVal = 1  # 将数据标签显示为数据对应的标签
# 设置标签位置，可选参数{'inBase', 'inEnd', 'l', 't', 'bestFit', 'outEnd', 'b', 'r', 'ctr'}
c1.dLbls.position = 't'  # 标签居上

c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=13)
c1.set_categories(dates)

# 设置系列1格式
s2 = c1.series[0]
# 设置线条颜色，不设置时默认为砖红色,其他颜色代码可参见 https://www.fontke.com/tool/rgb/00aaaa/
s2.graphicalProperties.line.solidFill = "FF5555"
s2.graphicalProperties.line.width = 30000  # 控制线条粗细
s2.smooth = True  # 设置为平滑曲线

# 插入图表
ws.add_chart(c1, "F2")

# 创建汇率折线图---美元欧元
# 数据区域为第2-3列，第2-13行，其他可自行调整
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=13)
c2 = LineChart()
c2.title = "美元及欧元汇率推移图"
c2.style = 12
# 设置为不显示图例，默认为显示图例
# c2.legend=None
# 设置图例居上
c2.legend.position = 't'
# 坐标轴设置
c2.y_axis.title = "汇率"
c2.y_axis.crossAx = 500
c2.y_axis.scaling.min = 5
c2.y_axis.scaling.max = 9
c2.y_axis.majorUnit = 1
c2.x_axis = DateAxis(crossAx=100)
# c2.x_axis.title = "月"

# 自定义图表的宽度和高度，不设置时默认为15*7.5cm
c2.width = 25
c2.height = 12

c2.dLbls = DataLabelList()
# c2.dLbls.showCatName = True  # 数据标签显示x坐标轴标签
c2.dLbls.showVal = 1  # 将数据标签显示为数据对应的标签
# 设置标签居上
c2.dLbls.position = 't'

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=13)
c2.set_categories(dates)

# 设置系列1格式
s1 = c2.series[0]
s1.graphicalProperties.line.width = 30000  # 控制线条粗细
s1.smooth = True  # 设置为平滑曲线

# 设置系列2格式
s2 = c2.series[1]
s2.graphicalProperties.line.solidFill = "00AAAA"  # 设置线条颜色
s2.graphicalProperties.line.width = 30000  # 控制线条粗细
s2.smooth = True  # 设置为平滑曲线

ws.add_chart(c2, "F28")

wb.save(save_dir)