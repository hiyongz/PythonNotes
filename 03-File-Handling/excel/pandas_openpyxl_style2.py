import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# https://blog.csdn.net/as604049322/article/details/111829106
# 设置pandas对象全部显⽰ 
pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)
import pandas as pd
from datetime import datetime, date
df = pd.DataFrame({'Date and time': [datetime(2015, 1, 1, 11, 30, 55),
                                     datetime(2015, 1, 2, 1, 20, 33),
                                     datetime(2015, 1, 3, 11, 10),
                                     datetime(2015, 1, 4, 16, 45, 35),
                                     datetime(2015, 1, 5, 12, 10, 15)],
                   'Dates only': [date(2015, 2, 1),
                                  date(2015, 2, 2),
                                  date(2015, 2, 3),
                                  date(2015, 2, 4),
                                  date(2015, 2, 5)],
                   'Numbers': [1010, 2020, 3030, 2020, 1515],
                   'Percentage': [.1, .2, .33, .25, .5],
                   })
df['final'] = [f"=C{i}*D{i}" for i in range(2, df.shape[0]+2)]


# df.to_excel("demo1.xlsx", sheet_name='Sheet1', index=False)
# 日期类型数据
# writer = pd.ExcelWriter("demo1.xlsx",
#                         datetime_format='mmm d yyyy hh:mm:ss',
#                         date_format='mmmm dd yyyy')
# df.to_excel(writer, sheet_name='Sheet1', index=False)
# writer.save()

# Styler对表格着色输出
df_style = df.style.applymap(lambda x: 'color:red', subset=["Date and time"]) \
    .applymap(lambda x: 'color:green', subset=["Dates only"]) \
    .applymap(lambda x: 'background-color:#ADD8E6', subset=["Numbers"]) \
    .background_gradient(cmap="PuBu", low=0, high=0.5, subset=["Percentage"])

writer = pd.ExcelWriter("demo_style.xlsx",
                        engine='openpyxl',
                        datetime_format='mmm d yyyy hh:mm:ss',
                        date_format='mmmm dd yyyy')
df_style.to_excel(writer, sheet_name='Sheet1', index=False)
writer.save()

# Pandas自适应列宽保存数据
#  计算表头的字符宽度
column_widths = (
    df.columns.to_series()
    .apply(lambda x: len(x.encode('gbk'))).values
)
#  计算每列的最大字符宽度
max_widths = (
    df.astype(str)
    .applymap(lambda x: len(x.encode('gbk')))
    .agg(max).values
)
# 计算整体最大宽度
widths = np.max([column_widths, max_widths], axis=0)

writer = pd.ExcelWriter("auto_column_width2.xlsx", engine='openpyxl')
df.to_excel(writer, sheet_name='Sheet1', index=False)
worksheet = writer.sheets['Sheet1']

for i, width in enumerate(widths, 1):
    worksheet.column_dimensions[get_column_letter(i)].width = width+1

writer.save()















