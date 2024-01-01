import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 设置pandas对象全部显⽰
pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)
def get_df():
    """
    ⽣成⽬标pandas对象
    ⽣成⼀个，shape是，（1２，２）的pandas对象，列索引是多重索引，其他三列是计算列，以后再加
    :return:
    """
    # axis=0索引
    indexs = [
    'aaa', 'bbb', 'ccc', 'ddd', 'eee', 'fff', 'ggg',
    'hhh', 'iii', 'jjj', 'kkk', 'lll'
    ]
    # axis = 1 是⼀个多层索引
    column_index_level_1 = ['CC'] * 2
    column_index_level_2 = ['AA', 'BB']
    # 多层索引，可以⽤多维数组创建
    column_all = [column_index_level_1, column_index_level_2]
    # 数据准备
    df = pd.DataFrame(data=np.random.rand(12, 2), index=indexs, columns=column_all)
    # 将数据扩⼤100万倍
    df = df * 100000
    return df


def add_column_and_row(df):
    """
    添加⾏和列，并且格式化百分位的数据
    :param df:
    :return:
    """
    # 添加 CC, 'AA-BB' 'AA+BB' 'AA/BB' 列
    df[('CC', 'AA-BB')] = df.apply(lambda x: x[('CC', 'AA')] - x[('CC', 'BB')], axis=1)
    df[('CC', 'AA+BB')] = df.apply(lambda x: x[('CC', 'AA')] + x[('CC', 'BB')], axis=1)
    df[('CC', 'AA/BB')] = df.apply(lambda x: x[('CC', 'AA')] / x[('CC', 'BB')] if x[('CC', 'BB')] != 0 else np.NAN, axis=1)
    # 添加 total⾏
    df.loc['total'] = df.apply(lambda x: x.sum(), axis=0)
    # AA/BB 格式化，并且保留两位⼩数
    df[('CC', 'AA/BB')] = df[('CC', 'AA/BB')].apply(lambda x: format(x, '.2%'))
def generate_excel(df, excel_name='D:/devWorkspace/PythonNotes/03-File-Handling/excel/test.xlsx', sheet_name='total'):
    """
    ⽣成excel
    :param sheet_name:
    :param excel_name:
    :param df:
    :return:
    """
    # 如果⽂件存在，直接添加sheet；不存在则创建新的⽂件
    excel_writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    if not os.path.exists(excel_name):
        df.to_excel(excel_writer, sheet_name=sheet_name)
    else:
        # book = openpyxl.Workbook(excel_name)
        book = openpyxl.load_workbook(excel_writer.path)
        excel_writer.book = book
        df.to_excel(excel_writer, sheet_name=sheet_name)
        excel_writer.save()
        excel_writer.close()

def set_excel_style(df, excel_name='./sample.xlsx', sheet_name='total'):
    book = openpyxl.load_workbook(excel_name)
    sheet = book[sheet_name]
    # 初始化字体样式
    font_bold = Font(name='Arial',
        size=8,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000',
        outline='None')
    font_not_bold = Font(name='Arial',
        size=8,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000',
        outline='None')
    # 插⼊前四⾏
    for i in range(1, 5):
        sheet.insert_rows(i)
    # 删除第7⾏默认的空⽩
    for row in range(7, sheet.max_row):
        for column in range(sheet.max_column):
            sheet[row][column].value = sheet[row + 1][column].value
    for cell in list(sheet.rows)[sheet.max_row - 1]:
        cell.value = None
    # 格式化数字
    for r, row in enumerate(sheet.rows):
        for c, cell in enumerate(row):
            if cell.row > 6:
                cell.number_format = '#,##0;-#,##0'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'), left=Side(), right=Side())
    # 设置整体的字体
    for row in sheet.rows:
        for cell in row:
            cell.font = font_not_bold
            cell.border = border_none
    # 前四⾏的第⼀个cell，加粗靠左
    sheet.cell(1, 1, value='Tittle').font = font_bold
    sheet.cell(2, 1, value='Period: Jun 2019').font = font_bold
    sheet.cell(3, 1, value='2019 Financial YTD Summary').font = font_bold
    sheet.cell(4, 1, value="(CNY '000)").font = font_bold
    for cell in sheet['A']:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if cell.value == 'total':
            cell.font = font_bold
            for cell__ in sheet[cell.row]:
                if cell__.column > 1:
                    cell__.border = border_double
                    cell__.font = font_bold
    # 合并列
    sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
    # 第五⾏ 第⼆列 设置背景颜⾊
    sheet.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(5, 2).fill = PatternFill(fill_type='solid', fgColor="B4C6E7")
    # 列折叠
    sheet.column_dimensions.group('B', 'D', hidden=True)
    # ⾏折叠
    sheet.row_dimensions.group(7, 7 + len(df.index) - 2, hidden=True)
    book.save(excel_name)

if __name__ == '__main__':
    df = get_df()
    add_column_and_row(df)
    generate_excel(df)
    # set_excel_style(df)  










